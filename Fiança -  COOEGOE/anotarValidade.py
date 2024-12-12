import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import os
from openpyxl import load_workbook
from marinetteSEFAZ import loginSEI, obterProcessosDeBloco,procurarArquivos, buscarInformacaoEmDocumento, escreverAnotacao, salvarPlanilha,buscarProcessoEmBloco
    
def encontrarFormaDePagamento():
    despachos = procurarArquivos(nav,"Despacho sobre Autorização de Despesa")
    regexBeneficiario = r"Beneficiário:(.*)\n"
    regexFormaDePagamento = r"Forma de Pagamento:(.*)\n"

    for despacho in reversed(despachos):
        try: 
            beneficiario = buscarInformacaoEmDocumento(nav,despacho,regexBeneficiario,"Rio de Janeiro").group(1).upper()
            formaPagamentoDespacho = buscarInformacaoEmDocumento(nav,despacho,regexFormaDePagamento,"Rio de Janeiro").group(1).upper()

        
            if "PERDIMENTO" in formaPagamentoDespacho:
                return "PERDIMENTO"

            if "BRADESCO" in formaPagamentoDespacho:
                return "Depósito Bradesco"
            
            if "CNPJ" in beneficiario:                      
                if any(palavra in formaPagamentoDespacho for palavra in ["GUIA", "GRERJ"]):
                    return "Guia"
                if any(palavra in formaPagamentoDespacho for palavra in ["GRU", "FUNAD"]):
                    return "Guia GRU"

            if "DEPÓSITO JUDICIAL" in formaPagamentoDespacho:
                return "Guia"

            regexBanco = r"(ITAÚ|Caixa Econômica|NUBANK|SANTANDER|Santander|C6 Bank|BANCO DO BRASIL|SICOOB|C6 BANK|PICPAY|CAIXA|CEF|SICRED|BANCO C6|MERCADO PAGO|ITAU)"
            banco = buscarInformacaoEmDocumento(nav,despacho,regexBanco,"Rio de Janeiro")
            if banco == None:
                banco = ""
            else:
                banco = banco.group(1)
            
    

            if "DEPÓSITO" in formaPagamentoDespacho:
                return "Depósito " + banco
            if  "CPF" in beneficiario:
                return "Depósito " + banco
            if "AGÊNCIA" in formaPagamentoDespacho:
                return "Depósito " + banco
        except:
            traceback.print_exc()
            pass
    
    try:
        indebitos = procurarArquivos(nav, "Correspondência Interna")
        formaDePagamento = buscarInformacaoEmDocumento(nav,indebitos[0],r"(INDÉBITO)").group(1)
        if formaDePagamento:
            return "Indébito"
    except:
        pass
    
    print("Não encontrado")
    return "ERRO"




def encontrarValidade():
    
    regexBanco = r"(BANCO DO BRASIL|GRERJ)"
    validade = "SEM VALIDADE"
    
    if tipo == "FIANÇA":
        docs = procurarArquivos(nav, ["Guia", "GRERJ"])
        for doc in reversed(docs):
            banco = buscarInformacaoEmDocumento(nav,doc,regexBanco)
            if banco != None:
                if banco.group(1) == "BANCO DO BRASIL":
                    regexValidade = r"(\d{2}\/\d{2}\/\d{4})\n"
                    
                elif banco.group(1) == "GRERJ":
                    regexValidade = r"(\d{2}\/\d{2}\/\d{4})"
                else:
                    continue
                
                validade = buscarInformacaoEmDocumento(nav,doc,regexValidade)
                print(validade)

                return validade.group(1)
                
        
        return validade

    if tipo =="EXECUÇÃO FISCAL":
        
        docs = procurarArquivos(nav, "Despacho sobre Autorização de Despesa")
        regexValidade = r"\bvalidade de (.*?)\b ? do referido"
        
        validade = buscarInformacaoEmDocumento(nav,docs[-1],regexValidade,)
               
        return validade.group(1)
    




marinette = r"C:\Users\jcampbell1\OneDrive - SEFAZ-RJ\CONTROLE GERENCIAL - PAGAMENTOS\Planilha Gerencial - Marinette.xlsx"


bloco = input("Digite o número do bloco: ")

wb = load_workbook(marinette)

if bloco not in wb.sheetnames:
    wb.create_sheet(bloco,0)
    wb.save(marinette)


tipo = int(input("Qual o tipo de bloco?\n1) Fiança\n2) Execução Fiscal\n3) Caução\n"))

match tipo:
    case 1:
        tipo = "FIANÇA"
    case 2:
        tipo = "EXECUÇÃO FISCAL"
    case 3:
        tipo = "CAUÇÃO"

nav = webdriver.Firefox()

loginSEI(nav, os.environ['login_sefaz'],os.environ['senha_sefaz'],"SEFAZ/COOEGOE")

processos = obterProcessosDeBloco(nav, bloco)

for i in range(1,len(processos)):
        
        WebDriverWait(nav,20).until(EC.invisibility_of_element_located(((By.XPATH, "//div[@class = 'sparkling-modal-close']"))))
        WebDriverWait(nav,20).until(EC.presence_of_element_located(((By.XPATH, "//tbody//tr"))))
        processo = nav.find_elements(By.XPATH, "//tbody//tr")[i]
        textoProcesso = processo.text
        
        linkProcesso = buscarProcessoEmBloco(nav,i)    
        nProcesso = linkProcesso.text
        
        if "FORMA DE PAGAMENTO" not in textoProcesso.upper():                          
            
            if tipo == "CAUÇÃO":

                continue
            
            validade = "-"
            
            if tipo == "EXECUÇÃO FISCAL":
                formaDePagamento = "DARJ"
            
            try:
                WebDriverWait(processo,20).until(EC.element_to_be_clickable(((By.XPATH, './/td[3]//a')))).click()
                nav.switch_to.window(nav.window_handles[1])
                
            
                if tipo == "FIANÇA":
                    formaDePagamento = encontrarFormaDePagamento()
    
                if "Depósito" not in formaDePagamento:
                    validade = encontrarValidade()
                
                
                texto = ["Forma de Pagamento: " + formaDePagamento]
                if validade != "-":
                    texto.append('Data de Validade da Guia: ' + validade)
            
            except:
                traceback.print_exc()
                continue
     
            finally:
                nav.close()
                nav.switch_to.window(nav.window_handles[0])
            
            try:
                escreverAnotacao(nav,texto,nProcesso)    
            except:
                traceback.print_exc()
                continue


nav.quit()
