import time
from time import sleep
import traceback
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import  Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import os
import re
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def login(navegador):
    navegador.get("https://sei.rj.gov.br/sip/login.php?sigla_orgao_sistema=ERJ&sigla_sistema=SEI")

    usuario = navegador.find_element(By.XPATH, value='//*[@id="txtUsuario"]')
    usuario.send_keys(os.environ['login_sefaz'])

    senha = navegador.find_element(By.XPATH, value='//*[@id="pwdSenha"]')
    senha.send_keys(os.environ['senha_sefaz'])

    exercicio = Select(navegador.find_element(By.XPATH, value='//*[@id="selOrgao"]'))
    exercicio.select_by_visible_text('SEFAZ')

    btnLogin = navegador.find_element(By.XPATH, value='//*[@id="Acessar"]')
    btnLogin.click()

    time.sleep(5)

    navegador.maximize_window()

    navegador.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)

def encontrarProcessos(navegador,blocoSolicitado,df,tipo):
    navegador.find_element(By.XPATH, "//span[text() = 'Blocos']").click()
    WebDriverWait(navegador,20).until(EC.element_to_be_clickable((By.XPATH, "//span[text() = 'Internos']"))).click()
    blocos = navegador.find_elements(By.XPATH, "//tbody//tr")[1:-1]

    for bloco in blocos:    
        nBloco = bloco.find_elements(By.XPATH,".//td")[1]
        if nBloco.text == blocoSolicitado:
            nBloco.find_element(By.XPATH, './/a').click()
            break
    processos = navegador.find_elements(By.XPATH, "//tbody//tr")
    time.sleep(1)

    for i in range(1,len(processos)):
        WebDriverWait(navegador,20).until(EC.invisibility_of_element_located(((By.XPATH, "//div[@class = 'sparkling-modal-close']"))))
        WebDriverWait(navegador,20).until(EC.presence_of_element_located(((By.XPATH, "//tbody//tr"))))
        processo = navegador.find_elements(By.XPATH, "//tbody//tr")[i]
        nProcesso = processo.find_element(By.XPATH, './/td[3]//a').text
        if nProcesso not in df['PROCESSO'].values:                          

            WebDriverWait(processo,20).until(EC.element_to_be_clickable(((By.XPATH, './/td[3]//a')))).click()

            time.sleep(3)
            if tipo == "FIANÇA":

                try:
                    
                    formaDePagamento = encontrarFormaDePagamento(navegador) 
                    validade = "-"
                    if formaDePagamento == "Guia GRU":
                        validade = "Sem Validade"
                    if formaDePagamento == "Guia":
                        print("Buscando Validade...")
                        navegador.switch_to.default_content()
                        validade,formaDePagamento = encontrarValidade(navegador, "FIANÇA")
                    df.loc[len(df)] = {"PROCESSO":nProcesso,"FORMA DE PAGAMENTO": formaDePagamento,"VALIDADE": validade}
                    salvarPlanilha(df)
                except:
                    traceback.print_exc()
                    pass
                finally:
                    navegador.close()
                    navegador.switch_to.window(navegador.window_handles[0])
                
            if tipo == "EXECUÇÃO FISCAL":
                try:
                    navegador.switch_to.window(navegador.window_handles[1])
                    validade,formaDePagamento = encontrarValidade(navegador, "EXECUÇÃO FISCAL")
                    print(formaDePagamento, validade)
                    navegador.switch_to.default_content()
 
                    df.loc[len(df)] = {"PROCESSO":nProcesso,"FORMA DE PAGAMENTO": formaDePagamento,"VALIDADE": validade}
                    salvarPlanilha(df)
                except:
                    traceback.print_exc()
                finally:
                    navegador.close()
                    navegador.switch_to.window(navegador.window_handles[0])
                try:
                    anotarFormaDePagamento(processo, formaDePagamento,navegador,validade)
                    navegador.switch_to.default_content()

                except:
                    traceback.print_exc()
            
                
        
        
def encontrarFormaDePagamento(navegador):
    navegador.switch_to.window(navegador.window_handles[1])
    WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvore")))
    docs = navegador.find_elements(By.XPATH, "//div[@id = 'divArvore']//div//a[@class = 'infraArvoreNo']")

    for doc in docs:
        try:
            docTexto = doc.text
            doc.click()

        except:
            pass
        if "Despacho sobre Autorização de Despesa" in docTexto:
            
            
            time.sleep(2)
            
            navegador.switch_to.default_content()            
            WebDriverWait(navegador,10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrVisualizacao")))
            WebDriverWait(navegador,10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvoreHtml")))
            try:
                beneficiario = navegador.find_element(By.XPATH, "//p//strong[contains(text(), 'Beneficiário')]" )
                beneficiario2 = navegador.find_element(By.XPATH, "//p[@class = 'Tabela_Texto_Alinhado_Esquerda' ][4]" )
                formaPagamentoDespacho = navegador.find_element(By.XPATH, "//p//strong[contains(text(), 'Forma de Pagamento')]" )
                forma2 =  navegador.find_element(By.XPATH, "//p[@class = 'Tabela_Texto_Alinhado_Esquerda' ][5]" )
                print(beneficiario.text)
                if "BRADESCO" in formaPagamentoDespacho.text.upper() or "BRADESCO" in forma2.text.upper():
                    formaPagamento = "Depósito Bradesco"
                    return formaPagamento
                if  "CPF" in beneficiario.text or "CPF" in beneficiario2.text:
                    formaPagamento = "Depósito"
                    
                    
                    
                    return formaPagamento

                if "CNPJ" in beneficiario.text or "CNPJ" in beneficiario2.text:
            
                
                    print(formaPagamentoDespacho.text)
                    
                    formaPagamento = ""
                    
                    if "GUIA" in formaPagamentoDespacho.text or "GUIA" in forma2.text:
                        formaPagamento = "Guia"
                    if "DEPÓSITO JUDICIAL" in formaPagamentoDespacho.text or "DEPÓSITO JUDICIAL" in forma2.text:
                        formaPagamento = "Guia"
                    elif "DEPÓSITO" in formaPagamentoDespacho.text or "DEPÓSITO" in forma2.text:
                        formaPagamento = "Depósito"
                    if "GRU" in formaPagamentoDespacho.text or "GRU" in forma2.text:
                        formaPagamento = "Guia GRU"
                    if "GRERJ" in formaPagamentoDespacho.text or "GRERJ" in forma2.text:
                        formaPagamento = "Guia"


              


                    return formaPagamento
            

                return ""
            except:
                traceback.print_exc()

def anotarFormaDePagamento(processo,formaPagamento,navegador,validade):

    print("Forma de Pagamento: " + formaPagamento)
    if validade != "-":
        print("Data de Validade da Guia: " + validade)
    

    processo.find_element(By.XPATH,".//td//a//img[@title='Anotações']").click()
                        
    time.sleep(2)
    try:
        iframe = navegador.find_element(By.TAG_NAME, 'iframe')
        navegador.switch_to.frame(iframe)

        txtarea = navegador.find_element(By.XPATH, '//textarea[@id = "txtAnotacao"]')

        txtarea.send_keys(Keys.PAGE_DOWN)
        txtarea.send_keys(Keys.END)
        txtarea.send_keys(Keys.ENTER)
        txtarea.send_keys("Forma de Pagamento: " + formaPagamento)
        if validade != "-":
            txtarea.send_keys(Keys.ENTER)
            txtarea.send_keys("Data de Validade da Guia: " + validade)
        time.sleep(1)
        salvar = navegador.find_element(By.XPATH, '//button[@value = "Salvar"]')

        salvar.click()
        
    except:
       traceback.print_exc()
       time.sleep(1)
       navegador.find_element(By.XPATH, "//div[@class = 'sparkling-modal-close']").click()
    finally:
        navegador.switch_to.default_content()

def encontrarValidade(navegador, tipo):
    navegador.switch_to.default_content()
    WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvore")))
    docs = navegador.find_elements(By.XPATH, "//div[@id = 'divArvore']//div//a[@class = 'infraArvoreNo']")
    quantDocs = len(docs)
    if tipo == "FIANÇA":
    
        for doc in range(quantDocs):
            docTexto = docs[doc].text
            if "Guia" in docTexto or "GRERJ" in docTexto:
                docs[doc].click()
                time.sleep(2)
                
                navegador.switch_to.default_content()            
                WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrVisualizacao")))
                WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvoreHtml")))
                spans = navegador.find_elements(By.XPATH, '//span')
                guia = ""
                for span in spans:
                    if "BANCO DO BRASIL" in span.get_attribute("innerHTML").upper():
                        guia = "Guia BB" 
                        break
                    if "GRERJ" in span.get_attribute("innerHTML").upper():
                        guia = "Guia GRERJ"
                        break
                    
                if guia == "Guia BB" or guia == "Guia GRERJ":
                    for span in spans:
                        #Regex pra achar as datas
                        regex = re.match("^\d{2}\/\d{2}\/\d{4}$",span.get_attribute('innerHTML'))
                        if regex:
                            validade = regex.group()
                            return validade,guia
                
                navegador.switch_to.default_content()            
                WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvore")))
                docs = navegador.find_elements(By.XPATH, "//div[@id = 'divArvore']//div//a[@class = 'infraArvoreNo']")    
        
        return "","Guia" 
    if tipo =="EXECUÇÃO FISCAL":
        validade= ""
        
        for doc in range(quantDocs):
            docTexto = docs[doc].text
            if "Despacho sobre Autorização de Despesa" in docTexto:
                try:
                    docs[doc].click()
                    time.sleep(2)
                    
                    navegador.switch_to.default_content()            
                    WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrVisualizacao")))
                    WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvoreHtml")))
                    
                    body = navegador.find_element(By.TAG_NAME, "body")
                    validade = re.search(r"\bvalidade de (.*?)\b do referido",body.text).group(1)
                    
                    
                except:
                    traceback.print_exc()
                finally:
                    navegador.switch_to.default_content()            
                    WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvore")))
                    docs = navegador.find_elements(By.XPATH, "//div[@id = 'divArvore']//div//a[@class = 'infraArvoreNo']")   
                
        return validade, "DARJ"

def salvarPlanilha(df):
    marinette = r"C:\Users\jcampbell1\OneDrive - SEFAZ-RJ\CONTROLE GERENCIAL - PAGAMENTOS\Planilha Gerencial - Marinette.xlsx"
    bloco = "869143"
    df = pd.read_excel(marinette, sheet_name=bloco)


    writer = pd.ExcelWriter(marinette, engine='openpyxl', mode='a', if_sheet_exists='replace')
    df.to_excel(writer, sheet_name=bloco, index=False)
    writer.close()

    planilha = load_workbook(marinette)
    tabela = planilha[bloco]
    for linha in range(2,tabela.max_row + 1):
        celula = tabela[f"D{linha}"]
        celula.value = f'=C{linha}-TODAY()'



    prazo = f'D2:D{tabela.max_row}'


    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

        
    red_rule = CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill)
        
    # Rule for cells with positive values (green)
    green_rule = CellIsRule(operator='greaterThan', formula=['15'], stopIfTrue=True, fill=green_fill)
    yellow_rule = CellIsRule(operator='lessThanOrEqual', formula=['15'], stopIfTrue=True, fill=yellow_fill)

    # Add rules to the worksheet
    tabela.conditional_formatting.add(prazo, red_rule)
    tabela.conditional_formatting.add(prazo, green_rule)


    planilha.save(marinette)
    planilha.close()



marinette = r"C:\Users\jcampbell1\OneDrive - SEFAZ-RJ\CONTROLE GERENCIAL - PAGAMENTOS\Planilha Gerencial - Marinette.xlsx"


bloco = input("Digite o número do bloco: ")

wb = load_workbook(marinette)

if bloco not in wb.sheetnames:
    wb.create_sheet(bloco,0)
    wb.save(marinette)

df = pd.read_excel(marinette, sheet_name=bloco)

try:
    print(df["PROCESSO"].values)

except:
    df = pd.DataFrame(columns=["PROCESSO", "FORMA DE PAGAMENTO", "VALIDADE",'PRAZO',"ACOMPANHAMENTO ESPECIAL","VALOR ORÇAMENTÁRIA", "VALOR EXTRAORÇAMENTÁRIA", "ERRO PAGAMENTO"], index=None)

tipo = int(input("Qual o tipo de bloco?\n1) Fiança\n2) Execução Fiscal\n"))

match tipo:
    case 1:
        tipo = "FIANÇA"
    case 2:
        tipo = "EXECUÇÃO FISCAL"


salvarPlanilha(df)

navegador = webdriver.Firefox()
login(navegador)

encontrarProcessos(navegador,bloco,df,tipo)
navegador.quit()