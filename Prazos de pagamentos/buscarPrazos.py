import time
from time import sleep
import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import  Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
from marinetteSEFAZ import salvarPlanilha

def login(user, password):
    navegador.get("https://siafe2.fazenda.rj.gov.br/Siafe/faces/login.jsp")
    usuario = WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="loginBox:itxUsuario::content"]')))
    usuario.send_keys(user)

    senha = navegador.find_element(By.XPATH, value='//*[@id="loginBox:itxSenhaAtual::content"]')
    senha.send_keys(password)

    exercicio = Select(navegador.find_element(By.XPATH, value='//*[@id="loginBox:cbxExercicio::content"]'))
    exercicio.select_by_visible_text('2024')

    btnLogin = navegador.find_element(By.XPATH, value='//*[@id="loginBox:btnConfirmar"]')
    btnLogin.click()
    navegador.maximize_window()
    
    navegador.get('https://siafe2.fazenda.rj.gov.br/Siafe/faces/execucao/financeira/ordemBancariaExtraOrcamentariaCad.jsp')
def popUp():
    try:
        WebDriverWait(navegador, 2).until(EC.element_to_be_clickable((By.XPATH,
        '//*[@id="pt1:warnMessageDec:newWarnMessagePopup::content"]//*[@id="pt1:warnMessageDec:frmExec:btnNewWarnMessageOK"]'))).click()
    except:
        None
def verificarSeOProcessoFoiPago(processo):
    WebDriverWait(navegador, 30).until(EC.element_to_be_clickable((By.XPATH, '// *[@id="pt1:tblOBExtra:sdtFilter::btn"]')))
    try:
        btnLimpar = navegador.find_element(By.XPATH, value= '//*[@id="pt1:tblOBExtra:btnClearFilter"]')
        btnLimpar.click()
    except:
        btnFiltro = navegador.find_element(By.XPATH, value='// *[@id="pt1:tblOBExtra:sdtFilter::disAcr"]')
        btnFiltro.click()
        try:
            btnLimpar = navegador.find_element(By.XPATH, value='//*[@id="pt1:tblOBExtra:btnClearFilter"]')
            btnLimpar.click()
        except:
            None
    
    WebDriverWait(navegador,20).until(EC.invisibility_of_element_located((By.XPATH, '//*[@id="pt1:tblOBExtra:table_rtfFilter:0:in_value_rtfFilter::content"]')))

    
    acessoRapido = navegador.find_element(By.ID, value='pt1:iTxtCad::content')

    acessoRapido.click()

    popUp()

    filtroProcessoSEI = Select(navegador.find_element(By.XPATH,
                                                        value='//*[@id="pt1:tblOBExtra:table_rtfFilter:0:cbx_col_sel_rtfFilter::content"]'))
    filtroProcessoSEI.select_by_visible_text('Processo')

    filtroContemProcesso = Select(navegador.find_element(By.XPATH,
                                                            value='//*[@id="pt1:tblOBExtra:table_rtfFilter:0:cbx_op_sel_rtfFilter::content"]'))
    filtroContemProcesso.select_by_visible_text('contém')
    valorFiltroProcesso = navegador.find_element(By.XPATH,
                                                    value='//*[@id="pt1:tblOBExtra:table_rtfFilter:0:in_value_rtfFilter::content"]')    
    valorFiltroProcesso.clear()
    valorFiltroProcesso.send_keys(processo)

    acessoRapido.click()
    
    WebDriverWait(navegador,15).until(EC.presence_of_element_located((By.XPATH, '//*[@id="pt1:tblOBExtra:table_rtfFilter:1:cbx_col_sel_rtfFilter::content"]')))
    
    
    
    try:
        WebDriverWait(navegador, 7).until( EC.element_to_be_clickable((By.XPATH, "//span[text()= '"+ processo + "']")) )
        if procurarErro():
            return "ERRO NO PAGAMENTO"
        return True
    except:
        try:
            navegador.find_element(By.XPATH, "//span[text()= '"+ processo.lower() + "']")
            if procurarErro():
                return "ERRO NO PAGAMENTO"
            return True
        except:
            try:
                navegador.find_element(By.XPATH, '//*[text() = "Não há dados para esta consulta."]')
                return False
            except:
                traceback.print_exc()
                return False
            
def procurarErro():
    tabelaDataResultado = navegador.find_element(By.XPATH, value='//*[@id="pt1:tblOBExtra:tabViewerDec::db"]')
    rows = tabelaDataResultado.find_elements(By.TAG_NAME, value="tr")
    if len(rows) > 0:
        for i in range(len(rows)):
            tabelaDataResultado = navegador.find_element(By.XPATH, value='//*[@id="pt1:tblOBExtra:tabViewerDec::db"]')
            rows = tabelaDataResultado.find_elements(By.TAG_NAME, value="tr")
            col = rows[i].find_elements(By.TAG_NAME, value="td")
            if col[12].text == "Erro no Pagamento":
                return True
            else:
                return False   

def preencherTabelaPrazos():
    planilhaMarinette = load_workbook(marinette)
    planilhaPrazos = load_workbook(caminhoPrazos)
    prazos = planilhaPrazos['PRAZOS']
    prazos.delete_rows(2,prazos.max_row)
    tabelas = planilhaMarinette.sheetnames
    x= 1
    celulasComPrazo = []
    for tabela in tabelas:
        if tabela == "PRAZOS":
            continue
        print(tabela)
        tabela = planilhaMarinette[tabela]
        x = 1
        for linha in tabela:
            prazo = tabela[f"D{x}"].value
            if "PAGO" not in prazo and  "PRAZO" not in prazo:
                try:
                    linhaAtual = []
                    
                    linhaAtual.append(tabela[f'A{x}'].value)
                    linhaAtual.append(tabela[f'B{x}'].value)
                    linhaAtual.append(tabela[f'C{x}'].value)                  
                    linhaAtual.append(tabela[f'D{x}'].value)                    

                    linhaAtual.append(tabela.title) 
                    
                    celulasComPrazo.append(linhaAtual)
                except:
                    traceback.print_exc()
            x += 1
        
    for linha in celulasComPrazo:
        try:
            print(linha)
            prazos.append(linha)
        except:
            traceback.print_exc()
    planilhaPrazos.save(caminhoPrazos)

    df = pd.read_excel(caminhoPrazos, sheet_name="PRAZOS")
    df = organizarPlanilha(df)
    salvarPlanilha(df,caminhoPrazos,"PRAZOS")

def organizarPlanilha(df):

    df['VALIDADE'] = df["VALIDADE"].apply(lambda x: datetime.strptime(x,"%d/%m/%Y").date() if pd.notnull(x) else None)

    df = df.sort_values(by = 'VALIDADE', na_position='first')

    df['VALIDADE'] = df["VALIDADE"].apply(lambda x: x.strftime("%d/%m/%Y") if pd.notnull(x) else None)

    return df


    
marinette = r"C:\Users\jcampbell1\OneDrive - SEFAZ-RJ\CONTROLE GERENCIAL - PAGAMENTOS\Planilha Gerencial - Marinette.xlsx"
planilha = load_workbook(marinette)
caminhoPrazos = r"C:\Users\jcampbell1\OneDrive - SEFAZ-RJ\CONTROLE GERENCIAL - PAGAMENTOS\Prazos de Guias - Marinette.xlsx"
tabelas = planilha.worksheets

navegador = webdriver.Firefox()
login(os.environ['cpf'], os.environ["senha_siafe"])

for tabela in tabelas[::-1]:
    
    bloco = tabela.title
    print(bloco)
    tabela = pd.read_excel(marinette, sheet_name= bloco, header=0)
    for index, linha in tabela.iterrows():  
        if "PAGO" not in str(linha['PRAZO']) and "PRAZO" not in str(linha['PRAZO']) :
            print(linha["PROCESSO"])
            try:
                resultado = verificarSeOProcessoFoiPago(linha['PROCESSO'])
                
                if resultado == False:
                    print("NAO PAGO")

                elif resultado == True:
                    print("TA PAGO")
                    tabela.loc[index,'VALIDADE'] = 'PAGO'
                    tabela.loc[index,'PRAZO'] = 'PAGO'
                elif resultado == "ERRO NO PAGAMENTO":
                    print(resultado)
                    tabela.loc[index,'PRAZO'] = resultado
                
                salvarPlanilha(tabela,marinette.bloco)
            except:
                traceback.print_exc()    
navegador.quit()
preencherTabelaPrazos()