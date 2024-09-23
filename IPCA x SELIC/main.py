import time
from time import sleep
import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import  Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import re
from openpyxl import load_workbook
from datetime import date
from glob import glob
from shutil import move
import pandas as pd


def loginSEI():
    navegador.get("https://sei.rj.gov.br/sip/login.php?sigla_orgao_sistema=ERJ&sigla_sistema=SEI")

    usuario = navegador.find_element(By.XPATH, value='//*[@id="txtUsuario"]')
    usuario.send_keys(os.environ['login_sefaz'])

    senha = navegador.find_element(By.XPATH, value='//*[@id="pwdSenha"]')
    senha.send_keys(os.environ['senha_sefaz'])

    exercicio = Select(navegador.find_element(By.XPATH, value='//*[@id="selOrgao"]'))
    exercicio.select_by_visible_text('SEFAZ')

    btnLogin = navegador.find_element(By.XPATH, value='//*[@id="Acessar"]')
    btnLogin.click()


    navegador.maximize_window()
    time.sleep(5)

    navegador.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
    
    trocarCoordenacao()
    
def trocarCoordenacao():
    coordenacao = navegador.find_elements(By.XPATH, "//a[@id = 'lnkInfraUnidade']")[1]
    print(coordenacao)
    if coordenacao.get_attribute("innerHTML") == 'SEFAZ/COOAJUR':
        print(coordenacao)
        coordenacao.click()
        WebDriverWait(navegador,5).until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'Trocar Unidade')]")))
        navegador.find_element(By.XPATH, "//td[text() = 'SEFAZ/COOEGOE' ]").click() 

def pesquisarProcesso(processo):
    navegador.switch_to.default_content()
    barraPesquisa = navegador.find_element(By.ID, "txtPesquisaRapida")

    barraPesquisa.send_keys(processo)
    barraPesquisa.send_keys(Keys.ENTER)
    
def abrirPastas():
    listaDocs = navegador.find_element(By.ID, "divArvore")
    pastas = listaDocs.find_elements(By.XPATH, '//a[contains(@id, "joinPASTA")]')
    for doc in pastas[:-1]:
        doc.click() 
        sleep(4) 
         
def procurarCalculo():

    time.sleep(3)    
    arvore = WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.ID, "ifrArvore")))    
    navegador.switch_to.frame(arvore)

    abrirPastas()

    listaDocs =  WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.ID, "divArvore")))  
    docs = listaDocs.find_elements(By.TAG_NAME, "a")

    for doc in reversed(docs):
        if "CÁLCULO" in doc.text.upper(): 
            doc.click()
            navegador.switch_to.default_content()            
            WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrVisualizacao")))
            WebDriverWait(navegador,20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "ifrArvoreHtml")))
            WebDriverWait(navegador,5).until(EC.presence_of_element_located((By.XPATH, '//*[contains(text(), "Calculadora do cidadão")]')))
            body = navegador.find_element(By.XPATH, '//body').text
            print(body)
            
            dataInicial = re.search(r"Data inicial\n (\d\d\/\d\d\d\d)", body).group(1)
            dataFinal = re.search(r"Data final\n (\d\d\/\d\d\d\d)", body).group(1)
            return dataInicial, dataFinal
            
        
def calcularSelic(dataInicial, dataFinal,valor):
    navegador.switch_to.new_window('tab')
    navegador.get("https://www3.bcb.gov.br/CALCIDADAO/publico/exibirFormCorrecaoValores.do?method=exibirFormCorrecaoValores&aba=4")
    WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.XPATH, '//input[@name = "dataInicial"]'))).send_keys(dataInicial)
    navegador.find_element(By.XPATH, '//input[@name = "dataFinal"]').send_keys(dataFinal)
    navegador.find_element(By.XPATH, '//input[@name = "valorCorrecao"]').send_keys(valor)
    navegador.find_element(By.XPATH, '//input[@title = "Corrigir valor"]').click()
    WebDriverWait(navegador,10).until(EC.presence_of_element_located((By.XPATH, '//input[@title = "Fazer nova pesquisa"]')))
    linhas = navegador.find_elements(By.XPATH, "//td")

    valor = re.search(r"(\d[\d\.]*,\d\d) ",linhas[-3].text).group(1)


    return valor

def salvarPlanilha(df,caminho):
    #SALVA A TABELA SEM APAGAR AS OUTRAS
    writer = pd.ExcelWriter(caminho, engine='openpyxl', mode='a', if_sheet_exists='replace')
    df.to_excel(writer, sheet_name="IPCAxSELIC", index=False)
    writer.close()

    planilha = load_workbook(caminho)
    tabela = planilha["IPCAxSELIC"]

    #FORMULA PARA PREENCHER A COLUNA DE PRAZO
    for linha in range(2,tabela.max_row + 1):
        celulaI = tabela[f"I{linha}"]
        celulaJ = tabela[f"J{linha}"]
        celulaI.value = f"=H{linha} - C{linha}"
        celulaJ.value = f"=I{linha} - D{linha}"

    #ALINHAR TAMANHO DAS CELULAS
    for column in tabela.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        tabela.column_dimensions[column_letter].width = adjusted_width

    planilha.save(caminho)
    planilha.close()





planilha = r"C:\Users\SEFAZ\Downloads\IPCAxSELIC.xlsx"
    
meses = ["","jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]




navegador = webdriver.Firefox()
loginSEI()

processos = pd.read_excel(planilha, dtype={"DATA DA ARRECADAÇÃO" :str,"DATA DO PGTO": str,"Principal": str } )

for i in range (len(processos)):
    processo = processos.iloc[i]
    if pd.isna(processo["Atualização SELIC"]):
        try:
        
            nProcesso = processo["Processo"]
            pesquisarProcesso(nProcesso)
            dataInicial, dataFinal = procurarCalculo()
        except:
            continue
        anoInicial = dataInicial.split("/")[1]
        anoFinal = dataFinal.split("/")[1]
        mesInicial = dataInicial.split("/")[0]
        mesFinal = dataFinal.split("/")[0]

        mesInicial = meses[int(mesInicial)]
        mesFinal = meses[int(mesFinal)]


        processos.loc[i,"DATA DA ARRECADAÇÃO"] = mesInicial + "/" + anoInicial
        processos.loc[i,"DATA DO PGTO"] = mesFinal + "/" + anoFinal
        try:
            processos.loc[i,"Atualização SELIC"] = calcularSelic("01" + dataInicial.replace("/", ""),"01" + dataFinal.replace("/", "") , processo["Principal"] + "00")
        except:
            continue

        salvarPlanilha(processos,planilha)

        navegador.close()
        navegador.switch_to.window(navegador.window_handles[0])



