import time
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import  Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import re
from openpyxl import load_workbook
from datetime import date
from glob import glob
from shutil import move
import tabula
from PyPDF2 import PdfReader

from marinetteSEFAZ import loginSEI,pesquisarProcesso, loginSIAFE, baixarArquivos

def alterarConfigSalvarPdf():
    
    navegador.get("about:preferences#general")

    pdf = WebDriverWait(navegador,5).until(EC.presence_of_element_located((By.XPATH,"//*[@type = 'application/pdf']" )))
    pdf.find_element(By.XPATH, ".//*[@class = 'actionsMenu']").click()
    pdf.find_element(By.XPATH,".//*[@label = 'Salvar arquivo']").click()
        
def baixarRelatorios():
    
    alterarConfigSalvarPdf()

    loginSEI(navegador, os.environ['login_sefaz'], os.environ['senha_sefaz'], "SEFAZ/COOEGOE")
    
    pesquisarProcesso(navegador, processoSEI)
    
    baixarArquivos(navegador,listaArquivos="FOLHA")
    
    alterarNomeArquivos()
    
    navegador.quit()

def verificarSaldo():
    loginSIAFE(navegador, os.environ['cpf'], os.environ["senha_siafe"])
    navegador.get(r"https://siafe2.fazenda.rj.gov.br/Siafe/faces/execucao/contabilidade/execucaoContabilidadeMain.jsp")

    WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.XPATH, "//a[text() = 'Detalhamento da Conta Contábil']"))).click()
    WebDriverWait(navegador,10).until(EC.element_to_be_clickable((By.XPATH, "//input[contains(@name, 'UnidadeGestora')]"))).send_keys("370200")

    select = Select(navegador.find_element(By.XPATH, '//select[@id = "tplSip:cbxMes::content"]'))
    
    if int(mes)> 12:
        x = 12
    else:
        x = mes
    
    select.select_by_index(int(x) + 1)
    navegador.find_element(By.XPATH, '//input[contains(@id, "SaldoZerado")]').click()
    time.sleep(2)
    credito = buscarSaldos("622110101", "2774")
    lme = buscarSaldos("823130201", "1.5.00.100.01")
    return credito,lme
     
def buscarSaldos(contabil,corrente):
   
    contaContabil = navegador.find_element(By.XPATH, '//input[contains(@id, "ContaContabil")]')

    contaContabil.clear()
    time.sleep(2)
    contaContabil.send_keys(contabil)
    contaContabil.send_keys(Keys.ENTER)
    time.sleep(2)
    contaCorrente = navegador.find_element(By.XPATH, '//input[contains(@id, "ContaCorrente")]')
    contaCorrente.clear()
    time.sleep(2)
    contaCorrente.send_keys(corrente)
    contaCorrente.send_keys(Keys.ENTER)
    time.sleep(2)
       
    tabela = navegador.find_element(By.XPATH, '//div[contains(@id, "tabListaFiltradaSaldo")]')
    contas = tabela.find_elements(By.XPATH, './/tr')
    
    saldos = []
    for conta in contas:
        celulas = conta.find_elements(By.XPATH, './/td')
        for i in range(0, len(celulas)):
            if i == 5:
                saldos.append(celulas[i].text)
    
    contaCorrente = navegador.find_element(By.XPATH, '//input[contains(@id, "ContaCorrente")]')
    for i in range (0, len(corrente)):
        contaCorrente.send_keys(Keys.BACKSPACE)
    time.sleep(2)
    contaContabil = navegador.find_element(By.XPATH, '//input[contains(@id, "ContaContabil")]')
    for i in range (0, len(contabil)):
        contaContabil.send_keys(Keys.BACKSPACE)
    
    print(saldos)
    return saldos

def alterarNomeArquivos():
    arquivo = ""
    downloads = r"C:\Users\\"+os.getlogin()+r"\Downloads"
    if not os.path.isdir(pasta):
        os.mkdir(pasta)
        
    while glob(downloads + "\\*.part") != []:
        time.sleep(0.1)
    
    arquivos = os.listdir(downloads)
    arquivos = [os.path.join(downloads, f) for f in arquivos]
    arquivo = max(arquivos, key=os.path.getmtime)
    print(arquivo)
    for processo in processos:
        if processo in arquivo.upper():
            nome = processo + "_" + meses[int(mes) - 1]
            newFile = pasta + "/" +  nome + ".pdf"
            move(arquivo, newFile)
            return
    os.remove(arquivo)

def somarValores(df,banco,index,colunaNome):
    return df[df[colunaNome].str.upper().str.contains(banco, case=False, na=False,)][index].astype(float).sum(numeric_only=True)

def atualizarMapaResumo():

    reader = PdfReader(tgrj0802p)
    page = reader.pages[0]
    text = page.extract_text()

    paragrafos = text.split("\n")

    valores = []
    for valor in paragrafos:
        if re.search(r",\d\d$", valor):
            valor = valor.split(" ")
            valor[1] = valor[1].replace('.', '')
            valor[1] = valor[1].replace(',', '.')
            valores.append(float(valor[1]))
            
    planilha = load_workbook(template)
    resumo = planilha["Mapa Resumo"]

    brutoServidores = resumo["C4"]
    brutoCotistas = resumo["C5"]
    descontos = resumo["C6"]        
            
    brutoServidores.value = valores[0]
    brutoCotistas.value = valores[1]
    descontos.value = valores[2]

    processoMapa = resumo["B2"]
    processoMapa.value = processoSEI
    
    competenciaMapa = resumo["C3"]
    competenciaMapa.value  = meses[int(mes) - 1] + "/" +  str(hoje.year)

    folhaDePagamentoMapa = resumo["B14"]
    folhaDePagamentoMapa.value = "Folha de Pagamento Encargos Gerais do Estado. Competência " + competenciaMapa.value + ". "+ processoMapa.value + "."

    dataExtracao = resumo["G15"]
    dataExtracao.value = hoje
    
    credito1 = resumo["H4"]
    credito1.value = credito[1]
    credito2 = resumo["H5"]
    credito2.value = credito[2]
    
    lme1 = resumo["H7"]
    lme1.value = lme[1]
    lme2 = resumo["H8"]
    lme2.value = lme[2]
    
    planilha.save(novaMemoria)

def atualizarSequencial():
    planilha = load_workbook(novaMemoria)
    sequencial = planilha["Sequencial"]
    retencao = planilha["Retenções"]

    celulas  ={
    '3190.11.23' :  sequencial["F36"],
    '3390.08.13' : sequencial['F40'],
    "3190.92.00": sequencial["F39"],
    "3190.11.34": sequencial["F38"],
    
    '3190.03.00': sequencial["M7"]
    }
    
    buscarValores(celulas,0,tgrj0807p,3,0)
    
    
    soma = 0
    
    for cell in range(5,32):
        valor = retencao[f'E{cell}'].value
        print(valor)
        if isinstance(valor, (float,int)):  # Verifique se o valor é um float
            soma += valor
            
    sequencial["M7"].value = sequencial["M7"].value - soma
    
    planilha.save(novaMemoria)

def atualizarRetencoes():
    planilha = load_workbook(novaMemoria)
    retencoes = planilha["Retenções"]

    celulasBancos = {
    "bvFinanceira": retencoes["E4"],
    "BANCO PAN": retencoes["E5"],
    "BANCO INDUSTRIAL": retencoes["E6"],
    "BMB": retencoes["E7"],
    "BANCO SANTANDER": retencoes["E8"],
    "BMG CARTAO": retencoes["E9"],
    "BANCO DAYCOVAL": retencoes["E10"],
    "caixa": retencoes["E11"],
    "4269": retencoes["E12"],
    "ccbb": retencoes["E13"],
    "BANCO BRADESCO": retencoes["E14"],
    "BANCO MASTER S.A": retencoes["E15"],
    "NIO MEIOS DE PAGAMENTO LTDA": retencoes["E16"],
    "BRADESCO FINANCIAMENTOS": retencoes["E17"],
    "BANCO ITAU CONSIGNADO S/A": retencoes["E18"],
    "bancoRs": retencoes["E19"],
    "FINANCEIRA": retencoes["E20"],
    "BANCO DO BRASIL": retencoes["E21"],
    "BENEFÍCIO CREDCESTA": retencoes["E22"],
    "BANCO INTER S.A.": retencoes["E23"],
    "proderj": retencoes["E24"],
    "repasseSefaz": retencoes["E25"]
}

    buscarValores(celulasBancos,0,pgov0832p,6,0)
    
    celulas = {
    "RIOPREVIDÊNCIA" : retencoes['E27'],
    "IMPOSTO DE RENDA": retencoes['E29'],
    "PENSÃO ALIMENTÍCIA": retencoes['E28']
    }
    
    buscarValores(celulas,-1,tgrj0801p,2,1)
    
    
    buscarValores({"RESSARCIMENTO FAZENDA ESTADUAL" : retencoes['E25']}, 2,tgrj0801p,5,1)
    

    proderj = retencoes["E24"]
    anulacoes=  retencoes["E31"]
    adiantamento = retencoes["E30"]
    encontrarRepasse(proderj)
    atualizarAdiantamento(anulacoes,0,5)
    atualizarAdiantamento(adiantamento,0,4)
    planilha.save(novaMemoria)

def buscarValores(dicionario,pagina,arquivo,colunaValores,colunaNome):

    df = tabula.read_pdf(arquivo, pages='all', pandas_options={'header': None})[pagina]
    print(df)
    df[colunaValores] = df[colunaValores].str.replace('.', '')
    df[colunaValores] = df[colunaValores].str.replace(',', '.')
    try:
        if arquivo == tgrj0801p and pagina == -1 :
            df[colunaValores + 1] = df[colunaValores +1].str.replace('.', '')
            df[colunaValores + 1] = df[colunaValores + 1].str.replace(',', '.')
            
            df[colunaValores] = df[colunaValores].fillna(0)
            df[colunaValores + 1] = df[colunaValores + 1].fillna(0) 
            
            df[colunaValores] = df[colunaValores].astype(float) + df[colunaValores + 1].astype(float)
    except:
        pass
    
    for chave,valor in dicionario.items():
        valor.value = somarValores(df,chave,colunaValores,colunaNome)

def atualizarAdiantamento(celula,index,pag):   
    reader = PdfReader( tgrj0801p)
    page = reader.pages[pag]
    text = page.extract_text()

    paragrafos = text.split("\n")

    valores = []
    for paragrafo in paragrafos:
        valor = re.findall(r"([\d|.]+,(\d\d))\b", paragrafo)
        if len(valor) > 1 :
            valores.append(valor[index][0])
           
    valores[0] = valores[0].replace(".", "")
    valores[0] = valores[0].replace(",", ".")
     
    celula.value = float(valores[0])
   
def encontrarRepasse(celula):

    reader = PdfReader(pgov0832p)
    page = reader.pages[1]
    text = page.extract_text()

    valorRepasse = re.search( r'([\d|.]+,(\d\d))\b \*\* Total de Repasse',text)
    valorRepasse = valorRepasse.group(1)
    
    valorRepasse = valorRepasse.replace(".", "")
    valorRepasse = valorRepasse.replace(",", ".")

    celula.value = float(valorRepasse)


def atualizarDecimoTerceiro():
    atualizarMapaResumo()

    
    planilha = load_workbook(novaMemoria)
    retencoes = planilha["Retenções"]
    sequencial = planilha["Sequencial"]

    celulas = {
    "PENSÃO ALIMENTÍCIA": retencoes['E28']
    }
    
    buscarValores(celulas,-1,tgrj0801p,6,0)

    celulas  ={
    '3190.11.25' :  sequencial["F37"],  
    }
    
    buscarValores(celulas,0,tgrj0807p,3,0)


    planilha.save(novaMemoria)
    

hoje = date.today()
meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", 
         "Dezembro","13","13_2"]

processos = ["PGOV0822P","PGOV0832P","TGRJ0801P","TGRJ0802P","TGRJ0807P", "TGRJ0808P"]
    
template = r"C:\Users\\"+ os.getlogin() +"\OneDrive - SEFAZ-RJ\Folha EGE\Exercício 2024\Teste - Marinete\Memória de Cálculo - Template.xlsx"

processoSEI = input("Digite o Processo: ")
mes = (input("Digite o mês da Folha: "))

navegador = webdriver.Firefox()


pasta = r"C:/Users/"+ os.getlogin() +r"/OneDrive - SEFAZ-RJ/Folha EGE/Exercício " + str(hoje.year) + '/' + str(mes) + "." + str(hoje.year) + ' - ' + processoSEI.replace("/", "_")
novaMemoria = pasta + "/" + "Memória de Cálculo - " + str(mes) + "." + str(hoje.year) + ".xlsx"

credito,lme = verificarSaldo()

# baixarRelatorios()

navegador.quit()

tgrj0807p = pasta + r"\TGRJ0807P_" + meses[int(mes) - 1] +".pdf"
pgov0832p = pasta + r"\PGOV0832P_" + meses[int(mes) - 1] +".pdf"
tgrj0802p = pasta + r"\TGRJ0802P_" + meses[int(mes) - 1] +".pdf"
tgrj0801p = pasta + r"\TGRJ0801P_" + meses[int(mes) - 1] +".pdf"

if int(mes) > 12:
    atualizarDecimoTerceiro()
    
else:
    if os.path.isfile(tgrj0801p) and os.path.isfile(tgrj0801p) and os.path.isfile(tgrj0801p) and os.path.isfile(tgrj0801p):
        atualizarMapaResumo()
        atualizarRetencoes()
        atualizarSequencial()

    else:
        print("Arquivos insuficientes!")