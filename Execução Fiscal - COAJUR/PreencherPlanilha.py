from marinetteSEFAZ import loginSEI, obterProcessosDeBloco, escreverAnotacao,buscarProcessoEmBloco, limparAnotacao,buscarInformacaoEmDocumento, procurarArquivos,incluirProcessoEmBloco,removerProcessoDoBloco
import time
import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import  Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
import os
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm
from datetime import datetime

def buscarNoDARJ():
        
    darj = procurarArquivos(nav, "DARJ")
    
    regexCDA = r"CERTIDÃO\n\n(.*)\n\n06"
    regexExecutado = r"NOME\n\n(.*\n? ?.*\n? ?.*\n? ?.*\n? ?.*\n? ?.*\n? ?.*\n? ?.*\n? ?.*)\n\n08"
    regexMontante = r"TOTAL A PAGAR\n\n(.*)\n\n14"
    lista = buscarInformacaoEmDocumento(nav,darj[-1],[regexCDA,regexExecutado,regexMontante],"DARJ",show=False)
        
    return lista[0].group(1), lista[1].group(1),lista[2].group(1)


def buscarProcessoJudicial():
    docs = procurarArquivos(nav,["Ofício", "Documento", "Petição", "Anexo Juntada", "Anexo SOLICITAÇÃO DA ANALISTA"])
    regexJudicial = r"\d{7}-\d{2}[\.\/]\d{4}\.\d\.?\d{2}\.\d{4}" 
    for doc in docs:
        pJudicial = buscarInformacaoEmDocumento(nav,doc,regexJudicial,["PROCURADORIA", "Justiça Estadual"],show=False)
        if pJudicial:
            return pJudicial.group()
    
    
    
    raise Exception("Processo Judicial não encontrado")
    
def preencherPlanilha(processo,nCDA,nomeExecutado,valorMontante,pJudicial,index):

    planilha = load_workbook(r"C:\Users\jcampbell1\Downloads\CONTROLE GERENCIAL - EXECUÇÃO FISCAL - COPIA.xlsx")

    sheet = planilha['EXECUÇÃO FISCAL']

    dataEntrada = sheet[f'B{index}']
    nProcesso = sheet[f'D{index}']
    materia = sheet[f'E{index}']
    cda = sheet[f'F{index}']
    executado = sheet[f'G{index}']
    montante = sheet[f'H{index}']
    procedimento = sheet[f'I{index}']
    status = sheet[f'J{index}']
    processoJudicial = sheet[f'K{index}']

    dataEntrada.value = hoje
    nProcesso.value = processo
    materia.value = "Execução Fiscal"
    cda.value = nCDA
    executado.value = nomeExecutado
    montante.value = valorMontante
    procedimento.value = "À COOEGOE PARA PAGAMENTO"
    status.value = "PENDENTE DE PGTO"
    processoJudicial.value = pJudicial
    
    planilha.save(r"C:\Users\jcampbell1\Downloads\CONTROLE GERENCIAL - EXECUÇÃO FISCAL - copia.xlsx")
  
  
def copiarPlanilha(caminhoOrigem,caminhoDestino):
    planilha = load_workbook(caminhoOrigem)
    planilha.save(caminhoDestino)  
        
nav = webdriver.Firefox()

hoje = datetime.today().strftime("%d/%m/%Y")
df = pd.read_excel(r"C:\Users\jcampbell1\Downloads\CONTROLE GERENCIAL - EXECUÇÃO FISCAL.xlsx", header=3)
index =  df['PROCESSO ADMINISTRATIVO'].last_valid_index()+ 5 + 1

caminhoOriginal = r"C:\Users\jcampbell1\Downloads\CONTROLE GERENCIAL - EXECUÇÃO FISCAL.xlsx"
caminhoCopia = r"C:\Users\jcampbell1\Downloads\CONTROLE GERENCIAL - EXECUÇÃO FISCAL - copia.xlsx"


copiarPlanilha(caminhoOriginal,caminhoCopia)

loginSEI(nav,os.environ['login_sefaz'], os.environ['senha_sefaz'], "SEFAZ/COOAJUR")

processos = obterProcessosDeBloco(nav,"938324")


#CRIAR EXCEÇÃO PARA PROCESSOS EM PROCESSO JUDICIAL 

try:
    for i in tqdm(range(1,len(processos[1:]) + 1)):
        processo = nav.find_elements(By.XPATH, "//tbody//tr")[i].text
        if "Processo inserido na Planilha" not in processo:
            linkProcesso = buscarProcessoEmBloco(nav,i)
            nProcesso = linkProcesso.text
            # limparAnotacao(nav,nProcesso)
            
            linkProcesso.click()
            print(nProcesso)
            nav.switch_to.window(nav.window_handles[1])
            try:
                cda,executado,montante = buscarNoDARJ()
                processoJudicial = buscarProcessoJudicial()
                preencherPlanilha(nProcesso,cda,executado,montante,processoJudicial,index)
                index += 1
            except:
                traceback.print_exc()
                continue
            finally:
                nav.close()
                nav.switch_to.window(nav.window_handles[0])
            
            try:
                escreverAnotacao(nav,["Processo inserido na Planilha"],nProcesso)
            except:
                traceback.print_exc()
        
        
        
except:
    traceback.print_exc()
    
copiarPlanilha(caminhoCopia,caminhoOriginal)
 
